from __future__ import print_function
from __future__ import division
from builtins import range
from past.utils import old_div
import os
import numpy as np
from collections import OrderedDict
import vtk
from vtk.util import numpy_support

def _checkDict(data, argname):
    if not isinstance(data, dict):
        raise "{:s} argument must be a dict of numpy arrays."
    default_keys = ("Coordinates", "Displacement", "Strain", "1st Principal Strain",
                    "2nd Principal Strain", "3rd Principal Strain", "Volumetric Strain")
    for k in default_keys:
        if k not in list(data.keys()):
            raise RuntimeError('{:s} dictionary should contain at least the following keys:\n'
                               '\t{:s}\n\t{:s}\n\t{:s}\n\t{:s}\n\t{:s}\n\t{:s}\n\t{:s}'.format(argname, *default_keys))

def vtkToDict(vtkgrid=None):
    """
    Convert vtkImageData to a data dictionary.

    Parameters
    ----------
    vtkgrid: vtkImageData
        VTK image grid with variables stored at vertices

    Returns
    -------
        data : dict
            Dictionary of numpy arrays of grid variables.
    """

    data = OrderedDict()
    data["Coordinates"] = np.zeros((vtkgrid.GetNumberOfPoints(), 3), float)
    for i in range(vtkgrid.GetNumberOfPoints()):
        vtkgrid.GetPoint(i, data["Coordinates"][i,:])

    for i in range(vtkgrid.GetPointData().GetNumberOfArrays()):
        arr = vtkgrid.GetPointData().GetAbstractArray(i)
        name = arr.GetName()
        if name != "Strain" and name != "Deformation Gradient":
            data[name] = numpy_support.vtk_to_numpy(arr)
        elif name == "Deformation Gradient":
            data[name] = np.transpose(numpy_support.vtk_to_numpy(arr).reshape(-1, 3, 3), axes=[0, 2, 1])
        else:
            data[name] = numpy_support.vtk_to_numpy(arr).reshape(-1, 3, 3)
    return data


def readNumpy(filename=None):
    """
    Load a data dictionary stored in Numpy npz format.

    Parameters
    ----------
    filename : str, required
        The path and filename of the file to load.

    Returns
    -------
    data : dict
        Dictionary of numpy arrays of grid variables.
    """
    contents = dict(np.load(filename))
    data = OrderedDict()
    # since numpy.savez_compressed does not respect OrderedDict() key order,
    # we have to rebuild a new OrderedDict()
    for k in ("Coordinates", "Deformation Gradient", "Displacement", "Strain", "1st Principal Strain",
              "2nd Principal Strain", "3rd Principal Strain", "Maximum Shear Strain",
              "Volumetric Strain"):
        try:
            data[k] = contents.pop(k)
        except:
            raise UserWarning("{:s} was not an entry in the data.".format(k))
    # additional keys beyond default are appended without control for order
    for k, v in list(contents.items()):
        data[k] = v
    return data

def readVTK(filename=None):
    reader = vtk.vtkXMLImageDataReader()
    reader.SetFileName(os.path.abspath(filename))
    reader.Update()
    vtkgrid = reader.GetOutput()

    data = OrderedDict()
    data["Coordinates"] = np.zeros((vtkgrid.GetNumberOfPoints(), 3), float)
    for i in range(vtkgrid.GetNumberOfPoints()):
        vtkgrid.GetPoint(i, data["Coordinates"][i,:])

    for i in range(vtkgrid.GetPointData().GetNumberOfArrays()):
        arr = vtkgrid.GetPointData().GetAbstractArray(i)
        name = arr.GetName()
        if arr.GetNumberOfComponents() != 9:
            data[name] = numpy_support.vtk_to_numpy(arr)
        elif name == "Deformation Gradient":
            data[name] = np.transpose(numpy_support.vtk_to_numpy(arr).reshape(-1, 3, 3), axes=[0, 2, 1])
        else:
            data[name] = numpy_support.vtk_to_numpy(arr).reshape(-1, 3, 3)
    return data

def readExcel(filename=None):
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    data = OrderedDict()
    for s in wb.sheetnames:
        ws = wb.get_sheet_by_name(s)
        if s == "Strain":
            #remove header and revert back to tensor notation
            data[s] = np.array(list(ws.values))[1:,[0, 3, 4, 3, 1, 5, 4, 5, 2]].reshape(-1, 3, 3).astype(float)
        elif s == "Deformation Gradient":
            data[s] = np.array(list(ws.values))[1:,:].reshape(-1, 3, 3).astype(float)
        else:
            data[s] = np.array(list(ws.values), float)
    return data

def extractRegion(data=None, irange=None, jrange=None, krange=None):
    """
    Extracts a subregion of grid and returns a data dictionary

    Parameters
    ----------
    irange : list, required
      start and end indices in the x dimension
    jrange : list, required
      start and end indices in the y dimension
    krange : list, required
      start and end indices in the z dimension

    Returns
    -------
      subregion : data dictionary
        A data dictionary for the gird subregion
    """
    grid = dictToVTK(data)
    extractFilter = vtk.vtkExtractVOI()
    extractFilter.SetInputData(grid)
    extractFilter.SetVOI(*irange, *jrange, *krange)
    extractFilter.Update()
    voi = extractFilter.GetOutput()
    subregion = vtkToDict(voi)
    return subregion

def calculateDifference(x=None, y=None, variable=None):
    """
    Calculates the difference between variables at all points.
    Order of difference is: *x* - *y*

    Parameters
    ----------
    x : dict, required
      Dictionary of ndarrays containing data to be compared.
    y : dict, required
      Dictionary of ndarrays containing data to be compared.

      .. note::
        *x* and *y* must have the same coordinates

    variable : str, required
      Dictionary key of variable to be compared.

    Returns
    -------
    differences : ndarray
      Absolute differences between variable values at all points.
    """
    _checkDict(x, "x")
    _checkDict(y, "y")

    if x["Coordinates"].shape[0] == y["Coordinates"].shape[0]:
        gridcheck = x["Coordinates"] - y["Coordinates"]
        gridcheck -= np.mean(gridcheck)
        if np.sum(gridcheck.ravel()) > 1e-7:
            raise RuntimeError("x and y dictionaries must have the same grid spacing.")
    else:
        raise RuntimeError("x and y dictionaries must have the same grid shape.")

    if variable != "Displacement" and len(x[variable].shape) == 2:
        if np.dot(x[variable][0,:], y[variable][0,:]) < 0.0:
            x[variable] = x[variable] * -1.0
        differences = x[variable] - y[variable]
    else:
        differences = x[variable] - y[variable]
    return differences

def calculateRMSDifference(x=None, y=None, variables=None):
    """
    Calculates the root-mean-square difference between variables in data dictionary.

    Parameters
    ----------
    x : dict, required
      Dictionary of ndarrays containing data to be compared.
    y : dict, required
      Dictionary of ndarrays containing data to be compared.

      .. note::
        *x* and *y* must be of the same shape.

    variables : [str,], required
      Dictionary keys of variables to be compared.

    Returns
    -------
    rmsd : dict
      Dictionary of root-mean-square differences for indicated variables.
    """
    rmsd = OrderedDict()
    for v in variables:
        if v != "Displacement" and len(x[v].shape) == 2:
            rmsd[v] = np.sqrt(old_div(np.sum(
                (np.linalg.norm(x[v], axis=1) - np.linalg.norm(y[v], axis=1)) ** 2), x[v].shape[0]))
        else:
            rmsd[v] = np.sqrt(old_div(np.sum((x[v].ravel() - y[v].ravel()) ** 2), x[v].size))
    return rmsd

def calculateStrainRatio(data=None, appliedDeformationGradient=None, value='33'):
    """
    Calculate the ratio between the applied strain and the indicated strain value.

    Parameters
    ----------
    data : dict
      Data dictionary with at least default items.
    appliedDeformationGradient : ndarray(3,3,float)
      The applied nominal strain.
    value : str
      The strain value to compare
       - '11' normal x strain
       - '22' normal y strain
       - '33' normal z strain
       - '12' xy shear strain
       - '13' xz shear strain
       - '23' yz shear strain
       - 'P1' the 1st principal strain
       - 'P2' the 2nd principal strain
       - 'P3' the 3rd principal strain

    Returns
    -------
    strain_ratio : ndarray(N)
      The strain ratio at N grid points.
    """
    _checkDict(data, "data")
    if appliedDeformationGradient.shape != (3,3):
        raise ValueError("appliedDeformationGradient must be of shape (3,3)")
    J = np.linalg.det(appliedDeformationGradient)
    if J <= 0:
        raise ValueError(("The determinant of the provided appliedDeformationGradient was {:e}"
                          " violating positive-definiteness. Please correct...").format(J))

    appliedStrain = 0.5 * (np.dot(appliedDeformationGradient.T, appliedDeformationGradient) - np.eye(3))
    if value in ('11', '22', '33', '12', '13', '23'):
        ind = (int(value[0]) - 1, int(value[1]) - 1)
        if np.abs(appliedStrain[ind[0], ind[1]]) < 1e-5:
            print("Warning: The applied strain value for the ratio requested is near zero. Interpret the results with caution.")
        print(data["Strain"].shape)
        strain_ratio = old_div(data["Strain"][:, ind[0], ind[1]],
                               appliedStrain[ind[0], ind[1]])
    elif value in ('P1', 'P2', 'P3'):
        l, v = np.linalg.eigh(appliedStrain)
        if value == 'P1':
            strain_ratio = old_div(data["1st Principal Strain"], l[2])
        elif value == 'P2':
            strain_ratio = old_div(data["2nd Principal Strain"], l[1])
        else:
            strain_ratio = old_div(data["3rd Principal Strain"], l[0])
    else:
        raise ValueError(("The provided argument: value={:s} is not recognized."
                          " Please indicate one of the following:\n '11'\n '22'\n '33'\n"
                          " '12'\n '13'\n '23'\n 'P1'\n 'P2'\n 'P3'").format(value))
    return strain_ratio

def dictToVTK(data=None):
    """
    Convert data dictionary to a vtkImageData grid.

    Parameters
    ----------
    data: dict
        Data dictionary with at least default items.

    Returns
    -------
        vtkgrid : vtkImageData
          Data variables stored on a vtk grid.
    """
    _checkDict(data, "data")
    vtkgrid = vtk.vtkImageData()

    tmp = np.sort(data["Coordinates"], axis=0)
    unique_partitions = [np.unique(tmp[:,i]) for i in range(3)]
    spacing = [unique_partitions[i][1] - unique_partitions[i][0] for i in range(3)]
    size = [unique_partitions[i].size for i in range(3)]
    vtkgrid.SetOrigin(np.min(data["Coordinates"], axis=0))
    vtkgrid.SetSpacing(spacing)
    vtkgrid.SetDimensions(size)
    for k, v in list(data.items()):
        if k == "Coordinates":
            continue
        else:
            if k == "Deformation Gradient":
                arr = numpy_support.numpy_to_vtk(np.transpose(v, axes=[0, 2, 1]).ravel(), deep=True, array_type=vtk.VTK_FLOAT)
            else:
                arr = numpy_support.numpy_to_vtk(v.ravel(), deep=True, array_type=vtk.VTK_FLOAT)
            arr.SetName(k)
            arr.SetNumberOfComponents(old_div(v.ravel().size, v.shape[0]))
            vtkgrid.GetPointData().AddArray(arr)
    return vtkgrid

def writeAsVTK(data, name=None):
    """
    Write *data* to disk in vtkImageData XML format.

    Parameters
    ----------
    data : dict, required
        The data dictionary containing at least default items.
    name : name, required
        Name of file to save to disk without the file suffix.
    """
    writer = vtk.vtkXMLImageDataWriter()
    if name is None:
        raise ("ERROR: Please provide a filename.")
    writer.SetFileName("{:s}.vti".format(name))
    if isinstance(data, dict):
        writer.SetInputData(dictToVTK(data))
    else:
        writer.SetInputData(data)
    writer.Update()
    writer.Write()
    print("... Wrote grid data to {:s}.vti".format(name))

def writeAsNumpy(data, name):
    """
    Save *data* to disk as an .npz file, which is an uncompressed zipped archive of
    multiple numpy arrays in .npy format. This can easily be loaded into memory from disk
    using numpy.load(filename).

    Parameters
    ----------
    data : dict, required
        Data dictionary with at least default items.
    name : str, required
        Name of file to save to disk without the file suffix.
    """
    print("... Saving file as numpy archive {:s}.npz".format(name))
    np.savez("{:s}.npz".format(name), **self.results)

def writeAsExcel(data, name):
    """
    Save *data* to disk as an xlsx file. Each key, value pair is saved on a separate sheet.

    Parameters
    ----------
    data : dict, required
        Data dictionary with at least default items.
    name : str, required
        Name of file to save to disk without the file suffix
    """
    from openpyxl import Workbook
    print("... Saving Results to {:s}.xlsx".format(name))
    wb = Workbook()
    ws = []
    for i, (k, v) in enumerate(data.items()):
        if i == 0:
            ws.append(wb.active)
            ws[-1].title = k
        else:
            ws.append(wb.create_sheet(title=k))
        if k == "Deformation Gradient":
            ws[i].append(["11", "12", "13", "21", "22", "23", "31", "32", "33"])
            d = v.reshape(-1, 9)
        elif len(v.shape) == 3:
            ws[i].append(["XX", "YY", "ZZ", "XY", "XZ", "YZ"])
            d = v.reshape(-1, 9)[:, [0, 4, 8, 1, 2, 5]]
        else:
            d = v
        if len(d.shape) > 1:
            for j in range(d.shape[0]):
                ws[i].append(list(d[j, :]))
        else:
            for j in range(d.shape[0]):
                ws[i].append([d[j]])

    wb.save(filename="{:s}.xlsx".format(name))
