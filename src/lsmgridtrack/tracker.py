from __future__ import print_function
from __future__ import division
from builtins import range
from builtins import object
from past.utils import old_div
import re
import fnmatch
import os
from collections import MutableMapping, OrderedDict
import yaml
import SimpleITK as sitk
import numpy as np
import vtk
from vtk.util import numpy_support
from openpyxl import Workbook

class FixedDict(MutableMapping):
    def __init__(self, data):
        self.__data = data

    def __len__(self):
        return len(self.__data)

    def __iter__(self):
        return iter(self.__data)

    def __setitem__(self, k, v):
        if k not in self.__data:
            raise KeyError("{:s} is not an acceptable key.".format(k))

        self.__data[k] = v

    def __delitem__(self, k):
        raise NotImplementedError

    def __getitem__(self, k):
        if k not in self.__data:
            raise KeyError("{:s} is not an acceptable key.".format(k))
        return self.__data[k]

    def __contains__(self, k):
        return k in self.__data

    def __repr__(self):
        return repr(self.__data)

class tracker(object):
    """
    Performs deformable image registration of laser scanning microscopy images
    with a photobleached grid.

    Parameters
    ----------
    reference_path : str
        The path to either a TIFF stack or 3D Nifti (.nii) image in the reference configuration.

        .. note::
            If *ref_img* is also specified this will be overridden.

    deformed_path : str
        The path to either a TIFF stack or 3D Nifti (.nii) image in a deformed configuration.

        .. note::
            If *def_img* is also specified this will be overridden.

    ref_img : SimpleITK.Image()
        Reference image stored in memory rather than read from disk.

        .. note::
            If *reference_path* is also specified, *ref_img* will supercede.

    def_img : SimpleITK.Image()
        Deformed image stored in memory rather than read from disk.

        .. note::
            If *deformed_path* is also specified, *def_img* will supercede.

    config : str, optional
        The path to a configuration file in YAML format where analysis options are set. If this
        is not provided, default options will be set.

        .. note::
            "Grid" key, value pairs must be set prior to running *execute()*.

    options: dict, optional
        Alternatively to using a configuration file, *options* can be set by providing a dictionary of
        key, value pairs to change. Likewise, the *options* attribute of the object can be modified.

        .. note::
            *options* keys are immutable.

        .. note::
            If both *options* and *config* are provided during object instantiation, values in *config* will
            supercede those in *options*.

    Attributes
    ----------
    transform : SimpleITK.Transform
        The composite transform calculated from the deformable registration.
    results : dict
        Results at N grid vertices

        * Coordinates - ndarray(N, 3) -- undeformed grid vertex locations
        * Displacement - ndarray(N, 3) -- grid vertex displacements
        * Deformation Gradient - ndarray(N, 3, 3) -- Deformation gradient at grid vertices
        * Strain - ndarray(N, 3, 3) -- Green-Lagrange strain tensors at grid vertices
        * 1st Principal Strain -- ndarray(N, 3) - 1st (Maximum) Principal Strain vectors at grid vertices
        * 2nd Principal Strain -- ndarray(N, 3) - 2nd (Median) Principal Strain vectors at grid vertices
        * 3rd Principal Strain -- ndarray(N, 3) - 3rd (Minimum) Principal Strain vectors at grid vertices
        * Maximum Shear Strain -- ndarray(N) - Maximum Shear Strain at grid vertices
        * Volumetric Strain -- ndarray(N) - Volumetric Strain at grid vertices
    vtkgrid : vtk.ImageData
        A VTK image of the tracked grid with displacements, deformation gradient, and strains stored at vertices.
    surface : vtk.vtkPolyData
        A polydata representation of the sample surface.
    """
    def __init__(self, **kwargs):
        # Default values
        self.options = FixedDict({
            "Image": FixedDict({
                "spacing": [1.0, 1.0, 1.0],
                "resampling": [1.0, 1.0, 1.0],
                "surface direction": "z_min"}),
            "Grid": FixedDict({
                "origin": False,
                "spacing": False,
                "size": False,
                "crop": False,
                "upsampling": 1}),
            "Registration": FixedDict({
                "method": "BFGS",
                "metric": "correlation",
                "iterations": 20,
                "sampling_fraction": 0.05,
                "sampling_strategy": 'RANDOM',
                "usemask": False,
                "reference landmarks": False,
                "deformed landmarks": False,
                "shrink_levels": [1],
                "sigma_levels": [0.0]})})

        self.reference_path = None
        self.deformed_path = None
        self.ref_img = None
        self.def_img = None

        self.config = None

        for key, value in kwargs.items():
            if key == "options":
                for k, v in value.items():
                    for k2, v2 in v.items():
                        self.options[k][k2] = v2
            else:
                setattr(self, key, value)

        if self.config is not None:
            self.parseConfig()

    def execute(self):
        """
        Executes the deformable image registration and post-analysis. """
        self._castOptions()
        self.results = OrderedDict()

        if self.ref_img is None and self.reference_path is not None:
            if self.options["Grid"]["crop"]:
                crop = self.options["Grid"]["origin"][2] + self.options["Grid"]["spacing"][2] * self.options["Grid"]["size"][2] + 2
            else:
                crop = self.options["Grid"]["crop"]
            self.ref_img = self.parseImg(self.reference_path, crop, self.options["Image"]["spacing"])
        if self.def_img is None and self.deformed_path is not None:
            self.def_img = self.parseImg(self.deformed_path, self.options["Grid"]["crop"], self.options["Image"]["spacing"])
        if np.abs(self.options["Image"]["resampling"] - 1.0).sum() > 1e-7:
            self.ref_img = self._resampleImage(self.ref_img, self.options["Image"]["resampling"])
            self.def_img = self._resampleImage(self.def_img, self.options["Image"]["resampling"])

        if self.options["Registration"]["usemask"]:
            self._makeMask()
        print("... Starting Deformable Registration")

        if np.any(self.options["Registration"]["reference landmarks"]) and np.any(self.options["Registration"]["deformed landmarks"]):
            #fixed_pts = self._defineFixedLandmarks()
            fixed_pts = (self.options["Registration"]["reference landmarks"] * np.array(self.ref_img.GetSpacing())).ravel()
            moving_pts = (self.options["Registration"]["deformed landmarks"] * np.array(self.ref_img.GetSpacing())).ravel()
            if moving_pts.size % 3 != 0:
                raise("ERROR: deformed image landmark index arrays must all be length 3.")
            if moving_pts.size != fixed_pts.size:
                raise "ERROR: {:d} deformed image landmarks were provided, while {:d} deformed were. Landmarks must correspond.".format(fixed_pts.size, moving_pts.size)
            # setup initial affine transform
            ix = sitk.BSplineTransformInitializer(self.ref_img, (3, 3, 3), 3)
            landmarkTx = sitk.LandmarkBasedTransformInitializerFilter()
            landmarkTx.SetFixedLandmarks(fixed_pts)
            landmarkTx.SetMovingLandmarks(moving_pts)
            landmarkTx.SetReferenceImage(self.ref_img)
            outTx = landmarkTx.Execute(ix)
        else:
            outTx = sitk.BSplineTransformInitializer(self.ref_img, (3, 3, 3), 3)

        rx = sitk.ImageRegistrationMethod()
        rx.AddCommand(sitk.sitkIterationEvent, lambda: self._printProgress(rx))
        if self.options["Registration"]["iterations"] > 0:
            print("... ... Finding optimal BSpline transform")
            rx.SetInitialTransform(outTx, True)
            if self.options["Registration"]["usemask"]:
                rx.SetMetricFixedMask(self._mask)
            if self.options["Registration"]["sampling_strategy"].upper() == "RANDOM":
                rx.SetMetricSamplingStrategy(rx.RANDOM)
                rx.SetMetricSamplingPercentagePerLevel(
                    (np.array(self.options["Registration"]["sampling_fraction"])*
                     np.array(self.options["Registration"]["shrink_levels"], dtype=float).tolist()))
            elif self.options["Registration"]["sampling_strategy"].upper() == "REGULAR":
                rx.SetMetricSamplingStrategy(rx.REGULAR)
                rx.SetMetricSamplingPercentagePerLevel(
                    (np.array(self.options["Registration"]["sampling_fraction"])*
                     np.array(self.options["Registration"]["shrink_levels"], dtype=float).tolist()), seed=31010)
            else:
                raise SystemError("Sampling strategy must be either: RANDOM or REGULAR")
            rx.SetInterpolator(sitk.sitkBSpline)
            if self.options["Registration"]["metric"] == "correlation":
                rx.SetMetricAsCorrelation()
            elif self.options["Registration"]["metric"] == "histogram":
                rx.SetMetricAsMattesMutualInformation()
            rx.SetMetricUseFixedImageGradientFilter(False)
            rx.SetShrinkFactorsPerLevel(self.options["Registration"]["shrink_levels"].tolist())
            rx.SetSmoothingSigmasPerLevel(self.options["Registration"]["sigma_levels"].tolist())
            if self.options["Registration"]["method"] == "ConjugateGradient":
                maximumStepSize = old_div(np.min(np.array(self.ref_img.GetSize()) * np.array(self.ref_img.GetSpacing())), 2.0)
                rx.SetOptimizerAsConjugateGradientLineSearch(1.0,
                                                             self.options["Registration"]["iterations"],
                                                             1e-5,
                                                             20,
                                                             lineSearchUpperLimit = 3.0,
                                                             maximumStepSizeInPhysicalUnits = maximumStepSize)
                rx.SetOptimizerScalesFromPhysicalShift()
            elif self.options["Registration"]["method"] == "GradientDescent":
                rx.SetOptimizerAsGradientDescent(1.0,
                                                 self.options["Registration"]["iterations"],
                                                 1e-5,
                                                 20,
                                                 rx.EachIteration)
                rx.SetOptimizerScalesFromPhysicalShift()
            elif self.options["Registration"]["method"] == "BFGS":
                rx.SetOptimizerAsLBFGSB(numberOfIterations = self.options["Registration"]["iterations"])
            outTx = rx.Execute(self.ref_img, self.def_img)
            print("... ... Optimal BSpline transform determined ")
            print("... ... ... Elapsed Iterations: {:d}\n... ... ... Final Metric Value: {:.5E}".format(rx.GetOptimizerIteration(),
                                                                                                        rx.GetMetricValue()))
        print("... Registration Complete")
        self.transform = outTx
        self.getGridDisplacements()
        self.getSampleSurface()
        self.getStrain()
        print("Analysis Complete!")

    def parseConfig(self):
        """
        Parse configuration file in YAML format.
        """
        with open(self.config) as user:
            user_settings = yaml.load(user, yaml.SafeLoader)

        for k, v in list(user_settings.items()):
            for k2, v2 in list(v.items()):
                self.options[k][k2] = v2

    def parseImg(self, p, crop, spacing):
        """
        Parse image from disk. Supports either a sequence of TIFF images or
        a NifTI (.nii) 3D image.
        """
        if p.endswith(".nii"):
            img = sitk.ReadImage(p, sitk.sitkFloat32)
            img.SetSpacing(spacing)
            arr = sitk.GetArrayFromImage(img)
            if crop:
                arr[:, :, crop:] = 0.0
            img2 = sitk.GetImageFromArray(arr)
            img2.CopyInformation(img)
            return sitk.RescaleIntensity(img2, 0.0, 1.0)

        files = fnmatch.filter(sorted(os.listdir(p)), "*.tif")
        counter = [re.search("[0-9]*\.tif", f).group() for f in files]
        for i, c in enumerate(counter):
            counter[i] = int(c.replace('.tif', ''))
        files = np.array(files, dtype=object)
        sorter = np.argsort(counter)
        files = files[sorter]
        img = []
        cnt = 0
        for fname in files:
            filename = os.path.join(p, fname)
            img.append(sitk.ReadImage(filename, sitk.sitkFloat32))
            if crop and cnt > crop:
                img[-1] *= 0.0
            cnt += 1
        img = sitk.JoinSeries(img)
        img.SetSpacing(spacing)
        print("\nImported 3D image stack ranging from {:s} to {:s}".format(files[0], files[-1]))
        return sitk.RescaleIntensity(img, 0.0, 1.0)

    def getGridDisplacements(self):
        r"""
        Using calculated transform from deformed image to reference image.
        """
        origin = (old_div(self.options["Grid"]["origin"], self.options["Image"]["resampling"])).astype(int)
        x = []
        for i in range(3):
            x.append(
                np.arange(origin[i] * self.ref_img.GetSpacing()[i],
                          (origin[i] + self.options["Grid"]["size"][i]*self.options["Grid"]["spacing"][i]) * self.ref_img.GetSpacing()[i],
                          self.options["Grid"]["spacing"][i] * self.ref_img.GetSpacing()[i] / float(self.options["Grid"]["upsampling"])))
        grid = np.meshgrid(x[0], x[1], x[2])
        self.results["Coordinates"] = np.zeros((grid[0].size, 3))
        self.results["Displacement"] = np.zeros((grid[0].size, 3))

        cnt = 0
        for k in range(grid[0].shape[2]):
            for i in range(grid[0].shape[0]):
                for j in range(grid[0].shape[1]):
                    p = np.array([grid[0][i,j,k],
                                  grid[1][i,j,k],
                                  grid[2][i,j,k]])
                    self.results["Coordinates"][cnt, :] = p
                    self.results["Displacement"][cnt, :] = self.transform.TransformPoint(p) - p
                    cnt += 1

    def getSampleSurface(self):
        """
        Finds the sample surface by using a line probe across z direction of the image and taking
        500 intensity samples. The first value along the line that exceeds 1/4 the mean intensity along
        entire line is considered the surface and that 3d point is stored in an array. This is repeated
        over a 28x28 grid. A surface is then constructed from the generated points. This is used in *getStrain()*
        to calculate the depth to each grid point as the minimum distance to the sample surface.

        Attributes
        ----------
        surface : vtk.vtkPolyData
          A polydata reprensentation of the sample surface
        """
        if self.options["Image"]["surface direction"] in ("x_min", "x_max"):
            ind = [1, 2]
        elif self.options["Image"]["surface direction"] in ("y_min", "y_max"):
            ind = [0, 2]
        else:
            ind = [0, 1]

        origin = list(self.ref_img.GetOrigin())
        spacing = list(self.ref_img.GetSpacing())
        size = list(self.ref_img.GetSize())

        u_coords = np.linspace(origin[ind[0]], origin[ind[0]] + spacing[ind[0]] * size[ind[0]], 30)[1:-1]
        w_coords = np.linspace(origin[ind[1]], origin[ind[1]] + spacing[ind[1]] * size[ind[1]], 30)[1:-1]

        vtk_image = sitk.SmoothingRecursiveGaussian(self.ref_img, 0.5)
        vtk_image = self._convertImageToVTK(vtk_image)

        probe = vtk.vtkProbeFilter()
        probe.SetSourceData(vtk_image)
        surface_points = vtk.vtkPoints()
        if self.options["Image"]["surface direction"] in ("x_max", "y_max", "z_max"):
            occurrence = -1
        else:
            occurrence = 0

        #normal direction
        ndir = [s for s in range(3) if s not in ind][0]
        for u in u_coords:
            for w in w_coords:
                line = vtk.vtkLineSource()
                p1 = np.zeros(3, dtype=float)
                p2 = np.zeros(3, dtype=float)
                p1[ind[0]] = u
                p1[ind[1]] = w
                p1[ndir] = origin[ndir]
                p2[ind[0]] = u
                p2[ind[1]] = w
                p2[ndir] = origin[ndir] + (spacing[ndir] * size[ndir])
                line.SetPoint1(*p1)
                line.SetPoint2(*p2)
                line.SetResolution(500)
                line.Update()
                probe.SetInputConnection(line.GetOutputPort())
                probe.Update()
                intensities = numpy_support.vtk_to_numpy(
                    probe.GetOutput().GetPointData().GetScalars())
                sind = np.argwhere(intensities > np.mean(intensities) / 4.0)[occurrence]
                pcoords = np.copy(p1)
                pcoords[ndir] = sind * (p2[ndir] - p1[ndir]) / 500.0
                surface_points.InsertNextPoint(pcoords)
        del vtk_image
        surfacePoly = vtk.vtkPolyData()
        surfacePoly.SetPoints(surface_points)

        reconstruct = vtk.vtkSurfaceReconstructionFilter()
        reconstruct.SetInputData(surfacePoly)
        reconstruct.Update()

        iso = vtk.vtkFlyingEdges3D()
        iso.SetInputData(reconstruct.GetOutput())
        iso.SetValue(0, 0.0)
        iso.Update()

        self.surface = iso.GetOutput()

    def getStrain(self):
        r"""
        Calculates the Green-Lagrange Strain at element centroids and interpolates to VTK
        grid vertices. Likewise, calculates principal and volumetric strains. For these values
        we need the deformation gradient, which assuming Einstein's summation convention unless explicitly
        indicated, follows as:

        .. note::
            Capital and lowercase letters imply reference and deformed configurations, respectively.

        Notation:

        .. math::

           F^i_{\,J} = \sum_{a=1}^{8} x^i_{\,a}\frac{\partial N_a}{\partial X^J}.

        We therefore need to determine :math:`\frac{\partial N_a}{\partial X^J}`.
        From the chain rule,

        .. math::

           \frac{\partial N_a}{\partial X^J} = \frac{\partial N_a}{\partial \eta^I} \left (\frac{\partial X^I}{\partial \eta^J} \right)^{-T}

        where

        .. math::

           \frac{\partial X^I}{\partial \eta^J} = \sum_{a=1}^{8} X^I_{\,a} \frac{\partial N_a}{\partial \eta^J}.

        The Green-Lagrange strain tensor then follows as,

        .. math::

           E_{IJ} = \frac{1}{2}\left(F_I^{\,i} g_{ij} F^j_{\,J} - G_{IJ}\right)

        where :math:`g_{ij}` is the spatial metric tensor and :math:`G_{IJ}` is the material metric tensor (both are the identity in Cartesian).

        The eigenvalues * eigenvectors of this tensor ordered decreasing by eigenvalue are the principal strains.
        The volumetric strain is,

        .. math::

           E_{volumetric} = \det{F^i_{\,J}} - 1.0.

        Returns
        -------
        vtkgrid : vtk.ImageData()
            VTK image grid with all results stored at grid vertices. 
        """
        vtkgrid = vtk.vtkImageData()
        vtkgrid.SetOrigin(np.array(self.options["Grid"]["origin"]) * np.array(self.ref_img.GetSpacing()))
        vtkgrid.SetSpacing(np.array(self.options["Grid"]["spacing"]) * np.array(self.ref_img.GetSpacing()) / float(self.options["Grid"]["upsampling"]))
        vtkgrid.SetDimensions(self.options["Grid"]["size"] * self.options["Grid"]["upsampling"])

        arr = numpy_support.numpy_to_vtk(self.results["Displacement"].ravel(), deep=True, array_type=vtk.VTK_DOUBLE)
        arr.SetNumberOfComponents(3)
        arr.SetName("Displacement")
        vtkgrid.GetPointData().SetVectors(arr)

        cells = vtkgrid.GetNumberOfCells()

        dNdEta = old_div(np.array([[-1, -1, -1],
                           [1, -1, -1],
                           [1, 1, -1],
                           [-1, 1, -1],
                           [-1, -1, 1],
                           [1, -1, 1],
                           [1, 1, 1],
                           [-1, 1, 1]], float), 8.0)

        Farray = np.zeros((cells, 3, 3), float)
        strain = np.zeros((cells, 3, 3), float)
        pstrain1 = np.zeros(cells, float)
        pstrain1_dir = np.zeros((cells, 3), float)
        pstrain2 = np.zeros(cells, float)
        pstrain2_dir = np.zeros((cells, 3), float)
        pstrain3 = np.zeros(cells, float)
        pstrain3_dir = np.zeros((cells, 3), float)
        vstrain = np.zeros(cells, float)
        maxshear = np.zeros(cells, float)
        order = [0, 1, 3, 2, 4, 5, 7, 6]
        for i in range(cells):
            nodeIDs = vtkgrid.GetCell(i).GetPointIds()
            X = numpy_support.vtk_to_numpy(vtkgrid.GetCell(i).GetPoints().GetData())
            X = X[order, :]
            x = np.zeros((8, 3), float)
            for j, k in enumerate(order):
                x[j, :] = X[j, :] + self.results["Displacement"][nodeIDs.GetId(k), :]
            dXdetaInvTrans = np.transpose(np.linalg.inv(np.einsum('ij,ik', X, dNdEta)))
            dNdX = np.einsum('ij,kj', dNdEta, dXdetaInvTrans)
            F = np.einsum('ij,ik', x, dNdX)
            Farray[i, :, :] = F
            C = np.dot(F.T, F)
            strain[i, :, :] = old_div((C - np.eye(3)), 2.0)
            l, v = np.linalg.eigh(strain[i, :, :])
            pstrain1[i] = l[2]
            pstrain2[i] = l[1]
            pstrain3[i] = l[0]
            pstrain1_dir[i, :] = v[:, 2]
            pstrain2_dir[i, :] = v[:, 1]
            pstrain3_dir[i, :] = v[:, 0]
            vstrain[i] = np.linalg.det(F) - 1.0
            maxshear[i] = np.abs(pstrain1[i] - pstrain3[i])
        for i in np.arange(1, pstrain1.shape[0]):
            if np.dot(pstrain1_dir[0,:], pstrain1_dir[i,:]) < 0:
                pstrain1_dir[i,:] *= -1.0
            if np.dot(pstrain2_dir[0,:], pstrain2_dir[i,:]) < 0:
                pstrain2_dir[i,:] *= -1.0
            if np.dot(pstrain3_dir[0,:], pstrain3_dir[i,:]) < 0:
                pstrain3_dir[i,:] *= -1.0

        def_grad = numpy_support.numpy_to_vtk(np.transpose(Farray, axes=[0, 2, 1]).ravel(), deep=True, array_type=vtk.VTK_FLOAT)
        def_grad.SetNumberOfComponents(9)
        def_grad.SetName("Deformation Gradient")

        vtk_strain = numpy_support.numpy_to_vtk(strain.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_strain.SetNumberOfComponents(9)
        vtk_strain.SetName("Strain")

        vtkgrid.GetCellData().SetTensors(vtk_strain)

        vtk_pstrain1 = numpy_support.numpy_to_vtk(pstrain1.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_pstrain1.SetNumberOfComponents(1)
        vtk_pstrain1.SetName("1st Principal Strain")

        vtk_pstrain1_dir = numpy_support.numpy_to_vtk(pstrain1_dir.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_pstrain1_dir.SetNumberOfComponents(3)
        vtk_pstrain1_dir.SetName("1st Principal Strain Direction")

        vtk_pstrain2 = numpy_support.numpy_to_vtk(pstrain2.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_pstrain2.SetNumberOfComponents(1)
        vtk_pstrain2.SetName("2nd Principal Strain")

        vtk_pstrain2_dir = numpy_support.numpy_to_vtk(pstrain2_dir.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_pstrain2_dir.SetNumberOfComponents(3)
        vtk_pstrain2_dir.SetName("2nd Principal Strain Direction")

        vtk_pstrain3 = numpy_support.numpy_to_vtk(pstrain3.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_pstrain3.SetNumberOfComponents(1)
        vtk_pstrain3.SetName("3rd Principal Strain")

        vtk_pstrain3_dir = numpy_support.numpy_to_vtk(pstrain3_dir.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_pstrain3_dir.SetNumberOfComponents(3)
        vtk_pstrain3_dir.SetName("3rd Principal Strain Direction")

        vtk_vstrain = numpy_support.numpy_to_vtk(vstrain.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_vstrain.SetNumberOfComponents(1)
        vtk_vstrain.SetName("Volumetric Strain")

        vtk_maxshear = numpy_support.numpy_to_vtk(maxshear.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_maxshear.SetNumberOfComponents(1)
        vtk_maxshear.SetName("Maximum Shear Strain")

        # depth from sample surface
        tree = vtk.vtkStaticPointLocator()
        tree.SetDataSet(self.surface)
        tree.BuildLocator()
        depth = np.zeros(self.results["Coordinates"].shape[0])
        for i in range(depth.size):
            p0 = self.results["Coordinates"][i,:]
            p1 = self.surface.GetPoint(tree.FindClosestPoint(p0))
            depth[i] = np.linalg.norm(np.array(p1)-p0)

        vtk_depth = numpy_support.numpy_to_vtk(depth.ravel(), deep=1, array_type=vtk.VTK_FLOAT)
        vtk_depth.SetNumberOfComponents(1)
        vtk_depth.SetName("Depth")

        vtkgrid.GetCellData().AddArray(def_grad)
        vtkgrid.GetCellData().AddArray(vtk_pstrain1)
        vtkgrid.GetCellData().AddArray(vtk_pstrain2)
        vtkgrid.GetCellData().AddArray(vtk_pstrain3)
        vtkgrid.GetCellData().AddArray(vtk_pstrain1_dir)
        vtkgrid.GetCellData().AddArray(vtk_pstrain2_dir)
        vtkgrid.GetCellData().AddArray(vtk_pstrain3_dir)
        vtkgrid.GetCellData().AddArray(vtk_vstrain)
        vtkgrid.GetCellData().AddArray(vtk_maxshear)
        vtkgrid.GetPointData().AddArray(vtk_depth)

        c2p = vtk.vtkCellDataToPointData()
        c2p.SetInputData(vtkgrid)
        c2p.Update()
        self.vtkgrid = c2p.GetOutput()
        names = ("Deformation Gradient", "Strain", "1st Principal Strain", "2nd Principal Strain",
                 "3rd Principal Strain", "1st Principal Strain Direction",
                 "2nd Principal Strain Direction", "3rd Principal Strain Direction",
                 "Maximum Shear Strain", "Volumetric Strain", "Depth")
        for a in names:
            if a != "Strain" and a != "Deformation Gradient":
                self.results[a] = numpy_support.vtk_to_numpy(self.vtkgrid.GetPointData().GetArray(a))
            elif a == "Deformation Gradient":
                self.results[a] = np.transpose(numpy_support.vtk_to_numpy(
                    self.vtkgrid.GetPointData().GetArray(a)).reshape(-1, 3, 3), axes=[0, 2, 1])
            else:
                self.results[a] = numpy_support.vtk_to_numpy(self.vtkgrid.GetPointData().GetArray(a)).reshape(-1, 3, 3)

    def _convertImageToVTK(self, img, sampling_factor=[1.0, 1.0, 1.0]):
        factor = self.options["Image"]["spacing"] / np.array(img.GetSpacing(), float) * np.array(sampling_factor) 
        img = self._resampleImage(img, factor)
        a = numpy_support.numpy_to_vtk(sitk.GetArrayFromImage(img).ravel(), deep=True, array_type=vtk.VTK_FLOAT)
        vtk_img = vtk.vtkImageData()
        vtk_img.SetOrigin(img.GetOrigin())
        vtk_img.SetSpacing(img.GetSpacing())
        vtk_img.SetDimensions(img.GetSize())
        vtk_img.GetPointData().SetScalars(a)
        return vtk_img

    def writeSurfaceAsVTK(self, name="surface"):
        if self.surface is None:
            raise AttributeError(("The surface has not been created yet. Either call *execute()",
                                  (" or explicitly call *getSampleSurface()")))
        writer = vtk.vtkXMLPolyDataWriter()
        writer.SetFileName('{}.vtp'.format(name))
        writer.SetInputData(self.surface)
        writer.Write()

    def writeImageAsVTK(self, img, name, sampling_factor=[1.0, 1.0, 1.0]):
        """
        Save image to disk as a .vti file. Image will be resampled such that it has spacing equal
        to that specified in *options["Image"]["spacing"]*.

        Parameters
        ----------
        img : SimpleITK.Image(), required
            Image in memory to save to disk. This is assumed to be either *ref_img* or *def_img*.
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving Image to {:s}.vti".format(name))
        vtk_img = self._convertImageToVTK(img)
        writer = vtk.vtkXMLImageDataWriter()
        writer.SetFileName("{:s}.vti".format(name))
        writer.SetInputData(vtk_img)
        writer.Write()

    def writeResultsAsVTK(self, name):
        """
        Save *vtkgrid* to disk as .vti file.

        Parameters
        ----------
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving Results to {:s}.vti".format(name))
        writer = vtk.vtkXMLImageDataWriter()
        writer.SetFileName("{:s}.vti".format(name))
        writer.SetInputData(self.vtkgrid)
        writer.Write()

    def writeResultsAsExcel(self, name):
        """
        Save *results* to disk as an xlsx file. Each key, value pair is saved on a separate sheet.

        Parameters
        ----------
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving Results to {:s}.xlsx".format(name))
        wb = Workbook()
        titles = ("Coordinates",
                  "Displacement",
                  "Deformation Gradient",
                  "Strain",
                  "1st Principal Strain",
                  "2nd Principal Strain",
                  "3rd Principal Strain",
                  "Maximum Shear Strain",
                  "Volumetric Strain",
                  "Depth")
        ws = []
        for i, t in enumerate(titles):
            if i == 0:
                ws.append(wb.active)
                ws[-1].title = t
            else:
                ws.append(wb.create_sheet(title=t))
            if t == "Strain":
                ws[i].append(["XX", "YY", "ZZ", "XY", "XZ", "YZ"])
                data = self.results[t].reshape(-1, 9)[:, [0, 4, 8, 1, 2, 5]]
            elif t == "Deformation Gradient":
                ws[i].append(["11", "12", "13", "21", "22", "23", "31", "32", "33"])
                data = self.results[t].reshape(-1, 9)
            elif "Principal" in t:
                data = self.results[t][:,np.newaxis] * self.results["{:s} Direction".format(t)]
            else:
                data = self.results[t]
            if len(data.shape) > 1:
                for j in range(data.shape[0]):
                    ws[i].append(list(data[j, :]))
            else:
                for j in range(data.shape[0]):
                    ws[i].append([data[j]])

        wb.save(filename="{:s}.xlsx".format(name))

    def writeResultsAsNumpy(self, name):
        """
        Save *results* to disk as an .npz file, which is an uncompressed zipped archive of
        multiple numpy arrays in .npy format. This can easily be loaded into memory from disk
        using numpy.load(filename).

        Parameters
        ----------
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving file as numpy archive {:s}.npz".format(name))
        np.savez_compressed("{:s}.npz".format(name), **self.results)

    def _resampleImage(self, img, factor):
        rs = sitk.ResampleImageFilter()
        rs.SetOutputOrigin(img.GetOrigin())
        rs.SetSize((np.array(factor) * np.array(img.GetSize())).astype(int).tolist())
        spacing = (old_div(np.array(img.GetSpacing()),
                           np.array(factor).astype(np.float32)).tolist())
        rs.SetOutputSpacing(spacing)
        rs.SetInterpolator(sitk.sitkLinear)
        return rs.Execute(img)

    def _makeMask(self):
        mask = (sitk.GetArrayFromImage(self.ref_img) * 0).astype(np.bool_)
        upper = self.options["Grid"]["origin"] + (self.options["Grid"]["size"] - 1) * self.options["Grid"]["spacing"] + 1
        mask[self.options["Grid"]["origin"][2]:upper[2], self.options["Grid"]["origin"][1]:upper[1], self.options["Grid"]["origin"][0]:upper[0]] = True
        self._mask = sitk.GetImageFromArray(mask.astype(np.uint8))
        self._mask.CopyInformation(self.ref_img)

    def _defineFixedLandmarks(self):
        edges = np.array(self.options["Grid"]["spacing"]) * np.array((self.options["Grid"]["size"] - 1))
        step = np.array([[0, 1, 0],
                        [1, 0, 0],
                        [0, -1, 0],
                        [-1, 0, 1],
                        [0, 1, 0],
                        [1, 0, 0],
                        [0, -1, 0]])
        fixedLandmarks = [self.options["Grid"]["origin"]]
        for i in range(step.shape[0]):
            fixedLandmarks.append(fixedLandmarks[-1] + edges * step[i,:])
        return (np.array(fixedLandmarks) * np.array(self.ref_img.GetSpacing())).ravel()

    def saveTransform(self, name):
        sitk.WriteTransform(self.transform, "{:s}.tfm".format(name))

    def _castOptions(self):
        arrays = (("Image", "spacing", "float"),
                  ("Image", "resampling", "float"),
                  ("Grid", "origin", "int"),
                  ("Grid", "spacing", "int"),
                  ("Grid", "size", "int"),
                  ("Grid", "upsampling", "int"),
                  ("Registration", "reference landmarks", "int"),
                  ("Registration", "deformed landmarks", "int"),
                  ("Registration", "shrink_levels", "int"),
                  ("Registration", "sigma_levels", "float"))
        for k1, k2, v in arrays:
            if k1 == "Grid" and k2 in ("spacing", "origin", "size"):
                if np.any(self.options[k1][k2]):
                    pass
                else:
                    raise SystemError("Values for Grid spacing, origin, and size must be provided before executing analysis.")
            if v == "float":
                self.options[k1][k2] = np.array(self.options[k1][k2], dtype=float)
            else:
                self.options[k1][k2] = np.array(self.options[k1][k2], dtype=int)

    def _printProgress(self, rx):
        print("... ... Elapsed Iterations: {:d}".format(rx.GetOptimizerIteration()))
        print("... ... Current Metric Value: {:.5E}".format(rx.GetMetricValue()))
