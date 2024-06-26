from __future__ import print_function
from __future__ import division
from builtins import range
from builtins import object
from past.utils import old_div
from collections import MutableMapping, OrderedDict
import yaml
import SimpleITK as sitk
import numpy as np
import vtkmodules.all as vtk
from vtkmodules.util import numpy_support
from openpyxl import Workbook
from argparse import ArgumentParser


class FixedDict(MutableMapping):
    def __init__(self, data):
        self.__data = data

    def __len__(self):
        return len(self.__data)

    def __iter__(self):
        return iter(self.__data)

    def __setitem__(self, k, v):
        if k not in self.__data:
            raise KeyError("{:s} is not an acceptable key.".format(k))

        self.__data[k] = v

    def __delitem__(self, k):
        raise NotImplementedError

    def __getitem__(self, k):
        if k not in self.__data:
            raise KeyError("{:s} is not an acceptable key.".format(k))
        return self.__data[k]

    def __contains__(self, k):
        return k in self.__data

    def __repr__(self):
        return repr(self.__data)


class tracker2d(object):
    """
    Performs deformable image registration of laser scanning microscopy images
    with a photobleached grid.

    Parameters
    ----------
    reference_path : str
        The path to either a TIFF stack or 3D Nifti (.nii) image in the reference configuration.

        .. note::
            If *ref_img* is also specified this will be overridden.

    deformed_path : str
        The path to either a TIFF stack or 3D Nifti (.nii) image in a deformed configuration.

        .. note::
            If *def_img* is also specified this will be overridden.

    ref_img : SimpleITK.Image()
        Reference image stored in memory rather than read from disk.

        .. note::
            If *reference_path* is also specified, *ref_img* will supercede.

    def_img : SimpleITK.Image()
        Deformed image stored in memory rather than read from disk.

        .. note::
            If *deformed_path* is also specified, *def_img* will supercede.

    config : str, optional
        The path to a configuration file in YAML format where analysis options are set. If this
        is not provided, default options will be set.

        .. note::
            "Grid" key, value pairs must be set prior to running *execute()*.

    options: dict, optional
        Alternatively to using a configuration file, *options* can be set by providing a dictionary of
        key, value pairs to change. Likewise, the *options* attribute of the object can be modified.

        .. note::
            *options* keys are immutable.

        .. note::
            If both *options* and *config* are provided during object instantiation, values in *config* will
            supercede those in *options*.

    Attributes
    ----------
    transform : SimpleITK.Transform
        The composite transform calculated from the deformable registration.
    results : dict
        Results at N grid vertices

        * Coordinates - ndarray(N, 2) -- undeformed grid vertex locations
        * Displacement - ndarray(N, 2) -- grid vertex displacements
        * Deformation Gradient - ndarray(N, 2, 2) -- Deformation gradient at grid vertices
        * Strain - ndarray(N, 2, 2) -- Green-Lagrange strain tensors at grid vertices
        * 1st Principal Strain -- ndarray(N, 2) - 1st (Maximum) Principal Strain vectors at grid vertices
        * 2nd Principal Strain -- ndarray(N, 2) - 2nd (Minimum) Principal Strain vectors at grid vertices
        * Maximum Shear Strain -- ndarray(N) - Maximum Shear Strain at grid vertices
        * Areal Strain -- ndarray(N) - Areal Strain at grid vertices
    vtkgrid : vtk.ImageData
        A VTK image of the tracked grid with displacements, deformation gradient, and strains stored at vertices.
    """

    def __init__(self, **kwargs):
        # Default values
        self.options = FixedDict(
            {
                "Image": FixedDict(
                    {
                        "spacing": [1.0, 1.0],
                        "resampling": [1.0, 1.0],
                    }
                ),
                "Grid": FixedDict(
                    {"origin": False, "spacing": False, "size": False, "upsampling": 1}
                ),
                "Registration": FixedDict(
                    {
                        "method": "BFGS",
                        "metric": "correlation",
                        "iterations": 20,
                        "sampling_fraction": 0.05,
                        "sampling_strategy": "RANDOM",
                        "reference landmarks": False,
                        "deformed landmarks": False,
                        "shrink_levels": [1],
                        "sigma_levels": [0.0],
                    }
                ),
            }
        )

        self.reference_path = None
        self.deformed_path = None
        self.ref_img = None
        self.def_img = None

        self.results = OrderedDict(
            {
                "Coordinates": None,
                "Displacement": None,
                "Deformation Gradient": None,
                "Strain": None,
                "1st Principal Strain": None,
                "2nd Principal Strain": None,
                "1st Principal Strain Direction": None,
                "2nd Principal Strain Direction": None,
                "Maximum Shear Strain": None,
                "Areal Strain": None,
            }
        )

        self.config = None

        for key, value in kwargs.items():
            if key == "options":
                for k, v in value.items():
                    for k2, v2 in v.items():
                        self.options[k][k2] = v2
            else:
                setattr(self, key, value)

        if self.config is not None:
            self.parseConfig()

    def execute(self):
        """
        Executes the deformable image registration and post-analysis."""
        self._castOptions()

        if self.ref_img is None and self.reference_path is not None:
            self.ref_img = self.parseImg(
                self.reference_path, self.options["Image"]["spacing"]
            )
        if self.def_img is None and self.deformed_path is not None:
            self.def_img = self.parseImg(
                self.deformed_path, self.options["Image"]["spacing"]
            )
        if np.abs(self.options["Image"]["resampling"] - 1.0).sum() > 1e-7:
            self.ref_img = self._resampleImage(
                self.ref_img, self.options["Image"]["resampling"]
            )
            self.def_img = self._resampleImage(
                self.def_img, self.options["Image"]["resampling"]
            )

        print("... Starting Deformable Registration")

        if np.any(self.options["Registration"]["reference landmarks"]) and np.any(
            self.options["Registration"]["deformed landmarks"]
        ):
            fixed_pts = (
                self.options["Registration"]["reference landmarks"]
                * np.array(self.ref_img.GetSpacing())
            ).ravel()
            moving_pts = (
                self.options["Registration"]["deformed landmarks"]
                * np.array(self.ref_img.GetSpacing())
            ).ravel()
            if moving_pts.size % 2 != 0:
                raise (
                    "ERROR: deformed image landmark index arrays must all be length 2."
                )
            if moving_pts.size != fixed_pts.size:
                raise "ERROR: {:d} deformed image landmarks were provided, while {:d} deformed were. Landmarks must correspond.".format(
                    fixed_pts.size, moving_pts.size
                )
            # setup initial affine transform
            ix = sitk.BSplineTransformInitializer(self.ref_img, (3, 3), 3)
            landmarkTx = sitk.LandmarkBasedTransformInitializerFilter()
            landmarkTx.SetFixedLandmarks(fixed_pts)
            landmarkTx.SetMovingLandmarks(moving_pts)
            landmarkTx.SetReferenceImage(self.ref_img)
            outTx = landmarkTx.Execute(ix)
        else:
            outTx = sitk.BSplineTransformInitializer(self.ref_img, (3, 3), 3)

        rx = sitk.ImageRegistrationMethod()
        rx.AddCommand(sitk.sitkIterationEvent, lambda: self._printProgress(rx))
        if self.options["Registration"]["iterations"] > 0:
            print("... ... Finding optimal BSpline transform")
            rx.SetInitialTransform(outTx, True)
            if self.options["Registration"]["sampling_strategy"].upper() == "RANDOM":
                rx.SetMetricSamplingStrategy(rx.RANDOM)
                rx.SetMetricSamplingPercentagePerLevel(
                    (
                        np.array(self.options["Registration"]["sampling_fraction"])
                        * np.array(
                            self.options["Registration"]["shrink_levels"], dtype=float
                        ).tolist()
                    )
                )
            elif self.options["Registration"]["sampling_strategy"].upper() == "REGULAR":
                rx.SetMetricSamplingStrategy(rx.REGULAR)
                rx.SetMetricSamplingPercentagePerLevel(
                    (
                        np.array(self.options["Registration"]["sampling_fraction"])
                        * np.array(
                            self.options["Registration"]["shrink_levels"], dtype=float
                        ).tolist()
                    ),
                    seed=31010,
                )
            else:
                raise SystemError("Sampling strategy must be either: RANDOM or REGULAR")
            rx.SetInterpolator(sitk.sitkBSpline)
            if self.options["Registration"]["metric"] == "correlation":
                rx.SetMetricAsCorrelation()
            elif self.options["Registration"]["metric"] == "histogram":
                rx.SetMetricAsMattesMutualInformation()
            rx.SetMetricUseFixedImageGradientFilter(False)
            rx.SetShrinkFactorsPerLevel(
                self.options["Registration"]["shrink_levels"].tolist()
            )
            rx.SetSmoothingSigmasPerLevel(
                self.options["Registration"]["sigma_levels"].tolist()
            )
            if self.options["Registration"]["method"] == "ConjugateGradient":
                maximumStepSize = old_div(
                    np.min(
                        np.array(self.ref_img.GetSize())
                        * np.array(self.ref_img.GetSpacing())
                    ),
                    2.0,
                )
                rx.SetOptimizerAsConjugateGradientLineSearch(
                    1.0,
                    self.options["Registration"]["iterations"],
                    1e-5,
                    20,
                    lineSearchUpperLimit=3.0,
                    maximumStepSizeInPhysicalUnits=maximumStepSize,
                )
                rx.SetOptimizerScalesFromPhysicalShift()
            elif self.options["Registration"]["method"] == "GradientDescent":
                rx.SetOptimizerAsGradientDescent(
                    1.0,
                    self.options["Registration"]["iterations"],
                    1e-5,
                    20,
                    rx.EachIteration,
                )
                rx.SetOptimizerScalesFromPhysicalShift()
            elif self.options["Registration"]["method"] == "BFGS":
                rx.SetOptimizerAsLBFGSB(
                    numberOfIterations=self.options["Registration"]["iterations"]
                )
            outTx = rx.Execute(self.ref_img, self.def_img)
            print("... ... Optimal BSpline transform determined ")
            print(
                "... ... ... Elapsed Iterations: {:d}\n... ... ... Final Metric Value: {:.5E}".format(
                    rx.GetOptimizerIteration(), rx.GetMetricValue()
                )
            )
        print("... Registration Complete")
        self.transform = outTx
        self.getGridDisplacements()
        self.getStrain()
        print("Analysis Complete!")

    def parseConfig(self):
        """
        Parse configuration file in YAML format.
        """
        with open(self.config) as user:
            user_settings = yaml.load(user, yaml.SafeLoader)

        for k, v in list(user_settings.items()):
            for k2, v2 in list(v.items()):
                self.options[k][k2] = v2

    def parseImg(self, p, spacing):
        """
        Parse TIFF image from disk.
        """
        img = sitk.ReadImage(p, sitk.sitkFloat32)
        img.SetSpacing(spacing)
        print(f"\nImported {p}")
        return sitk.RescaleIntensity(img, 0.0, 1.0)

    def getGridDisplacements(self):
        r"""
        Using calculated transform from deformed image to reference image.
        """
        origin = (
            old_div(self.options["Grid"]["origin"], self.options["Image"]["resampling"])
        ).astype(int)
        x = []
        for i in range(2):
            x.append(
                np.linspace(
                    origin[i] * self.ref_img.GetSpacing()[i],
                    (
                        origin[i]
                        + (self.options["Grid"]["size"][i] - 1)
                        * self.options["Grid"]["spacing"][i]
                    )
                    * self.ref_img.GetSpacing()[i],
                    self.options["Grid"]["size"][i] * self.options["Grid"]["upsampling"]
                    - (self.options["Grid"]["upsampling"] - 1),
                    endpoint=False,
                )
            )
        grid = np.meshgrid(x[0], x[1])
        self.results["Coordinates"] = np.zeros((grid[0].size, 2))
        self.results["Displacement"] = np.zeros((grid[0].size, 2))

        cnt = 0
        for i in range(grid[0].shape[0]):
            for j in range(grid[0].shape[1]):
                p = np.array([grid[0][i, j], grid[1][i, j]])
                self.results["Coordinates"][cnt, :] = p
                self.results["Displacement"][cnt, :] = (
                    self.transform.TransformPoint(p) - p
                )
                cnt += 1

    def getStrain(self):
        r"""
        Calculates the Green-Lagrange Strain at element centroids and interpolates to VTK
        grid vertices. Likewise, calculates principal and volumetric strains. For these values
        we need the deformation gradient, which assuming Einstein's summation convention unless explicitly
        indicated, follows as:

        .. note::
            Capital and lowercase letters imply reference and deformed configurations, respectively.

        Notation:

        .. math::

           F^i_{\,J} = \sum_{a=1}^{8} x^i_{\,a}\frac{\partial N_a}{\partial X^J}.

        We therefore need to determine :math:`\frac{\partial N_a}{\partial X^J}`.
        From the chain rule,

        .. math::

           \frac{\partial N_a}{\partial X^J} = \frac{\partial N_a}{\partial \eta^I} \left (\frac{\partial X^I}{\partial \eta^J} \right)^{-T}

        where

        .. math::

           \frac{\partial X^I}{\partial \eta^J} = \sum_{a=1}^{8} X^I_{\,a} \frac{\partial N_a}{\partial \eta^J}.

        The Green-Lagrange strain tensor then follows as,

        .. math::

           E_{IJ} = \frac{1}{2}\left(F_I^{\,i} g_{ij} F^j_{\,J} - G_{IJ}\right)

        where :math:`g_{ij}` is the spatial metric tensor and :math:`G_{IJ}` is the material metric tensor (both are the identity in Cartesian).

        The eigenvalues * eigenvectors of this tensor ordered decreasing by eigenvalue are the principal strains.
        The volumetric strain is,

        .. math::

           E_{areal} = \det{F^i_{\,J}} - 1.0.

        Returns
        -------
        vtkgrid : vtk.ImageData()
            VTK image grid with all results stored at grid vertices.
        """
        vtkgrid = vtk.vtkImageData()
        origin = [
            p / s
            for (p, s) in zip(self.options["Grid"]["origin"], self.ref_img.GetSpacing())
        ] + [0.0]
        spacing = [
            gs * s / float(self.options["Grid"]["upsampling"])
            for (gs, s) in zip(
                self.options["Grid"]["spacing"], self.ref_img.GetSpacing()
            )
        ] + [0.0]
        dimensions = self.options["Grid"]["size"] * self.options["Grid"][
            "upsampling"
        ] - (self.options["Grid"]["upsampling"] - 1)
        vtkgrid.SetOrigin(origin)
        vtkgrid.SetSpacing(spacing)
        vtkgrid.SetExtent(0, dimensions[0] - 1, 0, dimensions[1] - 1, 0, 0)

        displacements = np.concatenate(
            (
                self.results["Displacement"],
                np.zeros((self.results["Displacement"].shape[0], 1), dtype=float),
            ),
            axis=1,
        )
        arr = numpy_support.numpy_to_vtk(
            displacements.ravel(), deep=True, array_type=vtk.VTK_DOUBLE
        )
        arr.SetNumberOfComponents(3)
        arr.SetName("Displacement")
        vtkgrid.GetPointData().SetVectors(arr)

        cells = vtkgrid.GetNumberOfCells()

        dNdEta = np.array([[-1, -1], [1, -1], [1, 1], [-1, 1]], float) / 4.0

        Farray = np.zeros((cells, 3, 3), float)
        strain = np.zeros((cells, 3, 3), float)
        pstrain1 = np.zeros(cells, float)
        pstrain1_dir = np.zeros((cells, 3), float)
        pstrain2 = np.zeros(cells, float)
        pstrain2_dir = np.zeros((cells, 3), float)
        astrain = np.zeros(cells, float)
        maxshear = np.zeros(cells, float)
        order = [0, 1, 3, 2]
        for i in range(cells):
            nodeIDs = vtkgrid.GetCell(i).GetPointIds()
            X = numpy_support.vtk_to_numpy(vtkgrid.GetCell(i).GetPoints().GetData())
            X = X[order, 0:2]
            x = np.zeros((4, 2), float)
            for j, k in enumerate(order):
                x[j, :] = X[j, :] + self.results["Displacement"][nodeIDs.GetId(k), :]
            dXdetaInvTrans = np.transpose(np.linalg.inv(np.einsum("ij,ik", X, dNdEta)))
            dNdX = np.einsum("ij,kj", dNdEta, dXdetaInvTrans)
            F = np.einsum("ij,ik", x, dNdX)
            Farray[i, 0:2, 0:2] = F
            Farray[i, 2, 2] = 1.0
            C = np.dot(F.T, F)
            strain[i, 0:2, 0:2] = old_div((C - np.eye(2)), 2.0)
            strain[i, 2, 2] = 0.0
            l, v = np.linalg.eigh(strain[i, :, :])
            pstrain1[i] = l[1]
            pstrain2[i] = l[0]
            pstrain1_dir[i, :] = v[:, 1]
            pstrain2_dir[i, :] = v[:, 0]
            astrain[i] = np.linalg.det(F) - 1.0
            maxshear[i] = np.abs(pstrain1[i] - pstrain2[i])
        for i in np.arange(1, pstrain1.shape[0]):
            if np.dot(pstrain1_dir[0, :], pstrain1_dir[i, :]) < 0:
                pstrain1_dir[i, :] *= -1.0
            if np.dot(pstrain2_dir[0, :], pstrain2_dir[i, :]) < 0:
                pstrain2_dir[i, :] *= -1.0

        def_grad = numpy_support.numpy_to_vtk(
            np.transpose(Farray, axes=[0, 2, 1]).ravel(),
            deep=True,
            array_type=vtk.VTK_FLOAT,
        )
        def_grad.SetNumberOfComponents(9)
        def_grad.SetName("Deformation Gradient")

        vtk_strain = numpy_support.numpy_to_vtk(
            strain.ravel(), deep=1, array_type=vtk.VTK_FLOAT
        )
        vtk_strain.SetNumberOfComponents(9)
        vtk_strain.SetName("Strain")

        vtkgrid.GetCellData().SetTensors(vtk_strain)

        vtk_pstrain1 = numpy_support.numpy_to_vtk(
            pstrain1.ravel(), deep=1, array_type=vtk.VTK_FLOAT
        )
        vtk_pstrain1.SetNumberOfComponents(1)
        vtk_pstrain1.SetName("1st Principal Strain")

        vtk_pstrain1_dir = numpy_support.numpy_to_vtk(
            pstrain1_dir.ravel(), deep=1, array_type=vtk.VTK_FLOAT
        )
        vtk_pstrain1_dir.SetNumberOfComponents(3)
        vtk_pstrain1_dir.SetName("1st Principal Strain Direction")

        vtk_pstrain2 = numpy_support.numpy_to_vtk(
            pstrain2.ravel(), deep=1, array_type=vtk.VTK_FLOAT
        )
        vtk_pstrain2.SetNumberOfComponents(1)
        vtk_pstrain2.SetName("2nd Principal Strain")

        vtk_pstrain2_dir = numpy_support.numpy_to_vtk(
            pstrain2_dir.ravel(), deep=1, array_type=vtk.VTK_FLOAT
        )
        vtk_pstrain2_dir.SetNumberOfComponents(3)
        vtk_pstrain2_dir.SetName("2nd Principal Strain Direction")

        vtk_astrain = numpy_support.numpy_to_vtk(
            astrain.ravel(), deep=1, array_type=vtk.VTK_FLOAT
        )
        vtk_astrain.SetNumberOfComponents(1)
        vtk_astrain.SetName("Areal Strain")

        vtk_maxshear = numpy_support.numpy_to_vtk(
            maxshear.ravel(), deep=1, array_type=vtk.VTK_FLOAT
        )
        vtk_maxshear.SetNumberOfComponents(1)
        vtk_maxshear.SetName("Maximum Shear Strain")

        vtkgrid.GetCellData().AddArray(def_grad)
        vtkgrid.GetCellData().AddArray(vtk_strain)
        vtkgrid.GetCellData().AddArray(vtk_pstrain1)
        vtkgrid.GetCellData().AddArray(vtk_pstrain2)
        vtkgrid.GetCellData().AddArray(vtk_pstrain1_dir)
        vtkgrid.GetCellData().AddArray(vtk_pstrain2_dir)
        vtkgrid.GetCellData().AddArray(vtk_astrain)
        vtkgrid.GetCellData().AddArray(vtk_maxshear)

        c2p = vtk.vtkCellDataToPointData()
        c2p.SetInputData(vtkgrid)
        c2p.Update()
        self.vtkgrid = c2p.GetOutput()
        for a in self.results.keys():
            print(a)
            if a == "Displacement" or a == "Coordinates":
                continue
            elif a != "Strain" and a != "Deformation Gradient":
                self.results[a] = numpy_support.vtk_to_numpy(
                    self.vtkgrid.GetPointData().GetArray(a)
                )
            elif a == "Deformation Gradient":
                self.results[a] = np.transpose(
                    numpy_support.vtk_to_numpy(
                        self.vtkgrid.GetPointData().GetArray(a)
                    ).reshape(-1, 3, 3),
                    axes=[0, 2, 1],
                )
            else:
                self.results[a] = numpy_support.vtk_to_numpy(
                    self.vtkgrid.GetPointData().GetArray(a)
                ).reshape(-1, 3, 3)

    def _convertImageToVTK(self, img, sampling_factor=[1.0, 1.0]):
        factor = (
            self.options["Image"]["spacing"]
            / np.array(img.GetSpacing(), float)
            * np.array(sampling_factor)
        )
        img = self._resampleImage(img, factor)
        a = numpy_support.numpy_to_vtk(
            sitk.GetArrayFromImage(img).ravel(), deep=True, array_type=vtk.VTK_FLOAT
        )
        vtk_img = vtk.vtkImageData()
        vtk_img.SetOrigin(img.GetOrigin())
        vtk_img.SetSpacing(img.GetSpacing())
        vtk_img.SetDimensions(img.GetSize())
        vtk_img.GetPointData().SetScalars(a)
        return vtk_img

    def writeImageAsVTK(self, img, name):
        """
        Save image to disk as a .vti file. Image will be resampled such that it has spacing equal
        to that specified in *options["Image"]["spacing"]*.

        Parameters
        ----------
        img : SimpleITK.Image(), required
            Image in memory to save to disk. This is assumed to be either *ref_img* or *def_img*.
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving Image to {:s}.vti".format(name))
        if isinstance(img, vtk.vtkImageData):
            vtk_img = img
        else:
            vtk_img = self._convertImageToVTK(img)
        writer = vtk.vtkXMLImageDataWriter()
        writer.SetFileName("{:s}.vti".format(name))
        writer.SetInputData(vtk_img)
        writer.Write()

    def writeResultsAsVTK(self, name):
        """
        Save *vtkgrid* to disk as .vti file.

        Parameters
        ----------
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving Results to {:s}.vti".format(name))
        writer = vtk.vtkXMLImageDataWriter()
        writer.SetFileName("{:s}.vti".format(name))
        writer.SetInputData(self.vtkgrid)
        writer.Write()

    def writeResultsAsExcel(self, name):
        """
        Save *results* to disk as an xlsx file. Each key, value pair is saved on a separate sheet.

        Parameters
        ----------
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving Results to {:s}.xlsx".format(name))
        wb = Workbook()
        titles = self.results.keys()
        ws = []
        for i, t in enumerate(titles):
            if i == 0:
                ws.append(wb.active)
                ws[-1].title = t
            else:
                ws.append(wb.create_sheet(title=t))
            if t == "Strain":
                ws[i].append(["XX", "YY", "XY"])
                data = self.results[t].reshape(-1, 9)[:, [0, 4, 1]]
            elif t == "Deformation Gradient":
                ws[i].append(["11", "12", "21", "22"])
                data = self.results[t].reshape(-1, 9)[:, [0, 1, 3, 4]]
            elif "Principal" in t and not ("Direction" in t):
                data = (
                    self.results[t][:, np.newaxis]
                    * self.results["{:s} Direction".format(t)]
                )
            else:
                data = self.results[t]
            if len(data.shape) > 1:
                for j in range(data.shape[0]):
                    ws[i].append(list(data[j, :]))
            else:
                for j in range(data.shape[0]):
                    ws[i].append([data[j]])

        wb.save(filename="{:s}.xlsx".format(name))

    def writeResultsAsNumpy(self, name):
        """
        Save *results* to disk as an .npz file, which is an uncompressed zipped archive of
        multiple numpy arrays in .npy format. This can easily be loaded into memory from disk
        using numpy.load(filename).

        Parameters
        ----------
        name : str, required
            Name of file to save to disk without the file suffix
        """
        print("... Saving file as numpy archive {:s}.npz".format(name))
        np.savez_compressed("{:s}.npz".format(name), **self.results)

    def _resampleImage(self, img, factor):
        rs = sitk.ResampleImageFilter()
        rs.SetOutputOrigin(img.GetOrigin())
        rs.SetSize((np.array(factor) * np.array(img.GetSize())).astype(int).tolist())
        spacing = old_div(
            np.array(img.GetSpacing()), np.array(factor).astype(np.float32)
        ).tolist()
        rs.SetOutputSpacing(spacing)
        rs.SetInterpolator(sitk.sitkLinear)
        return rs.Execute(img)

    def saveTransform(self, name):
        sitk.WriteTransform(self.transform, "{:s}.tfm".format(name))

    def _castOptions(self):
        arrays = (
            ("Image", "spacing", "float"),
            ("Image", "resampling", "float"),
            ("Grid", "origin", "int"),
            ("Grid", "spacing", "int"),
            ("Grid", "size", "int"),
            ("Grid", "upsampling", "int"),
            ("Registration", "reference landmarks", "int"),
            ("Registration", "deformed landmarks", "int"),
            ("Registration", "shrink_levels", "int"),
            ("Registration", "sigma_levels", "float"),
        )
        for k1, k2, v in arrays:
            if k1 == "Grid" and k2 in ("spacing", "origin", "size"):
                if np.any(self.options[k1][k2]):
                    pass
                else:
                    raise SystemError(
                        "Values for Grid spacing, origin, and size must be provided before executing analysis."
                    )
            if v == "float":
                self.options[k1][k2] = np.array(self.options[k1][k2], dtype=float)
            else:
                self.options[k1][k2] = np.array(self.options[k1][k2], dtype=int)

    def _printProgress(self, rx):
        print("... ... Elapsed Iterations: {:d}".format(rx.GetOptimizerIteration()))
        print("... ... Current Metric Value: {:.5E}".format(rx.GetMetricValue()))


if __name__ == "__main__":
    parser = ArgumentParser(
        description="Perform a deformable image registration of provided images and configuration file."
    )
    parser.add_argument(
        "--configuration_file", type=str, nargs=1, help="Path to configuration file"
    )
    parser.add_argument(
        "--reference_path",
        type=str,
        nargs=1,
        help="Path to reference image file or directory containing stack of images.",
    )
    parser.add_argument(
        "--deformed_path",
        type=str,
        nargs=1,
        help="Path to deformed image file or directory containing stack of images.",
    )
    parser.add_argument("--vtk_out", type=str, nargs="?", default="output")
    parser.add_argument("--excel_out", type=str, nargs="?", default=None)

    args = parser.parse_args()

    tracker = tracker2d(
        config=args.configuration_file[0],
        reference_path=args.reference_path[0],
        deformed_path=args.deformed_path[0],
    )

    tracker.execute()
    tracker.writeResultsAsVTK(args.vtk_out)
    if args.excel_out:
        tracker.writeResultsAsExcel(args.excel_out)
